import openpyxl
import csv

wb = openpyxl.load_workbook('journal_pone_0139539_s005.xlsx')
#print wb.get_sheet_names()

#For bengali recipes
csvfile = open('bengali.csv', 'wb')
writer = csv.writer(csvfile)
bengali = wb.get_sheet_by_name('Bengali Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, bengali.max_row+1):
	if (recipe_no != bengali.cell(row=i, column=1).value):
		recipe_no = bengali.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == bengali.cell(row=i, column=1).value):
			ingredients.append(bengali.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For gujarati recipes
csvfile = open('gujarati.csv', 'wb')
writer = csv.writer(csvfile)
gujarati = wb.get_sheet_by_name('Gujarati Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, gujarati.max_row+1):
	if (recipe_no != gujarati.cell(row=i, column=1).value):
		recipe_no = gujarati.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == gujarati.cell(row=i, column=1).value):
			ingredients.append(gujarati.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For jain recipes
csvfile = open('jain.csv', 'wb')
writer = csv.writer(csvfile)
jain = wb.get_sheet_by_name('Jain Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, jain.max_row+1):
	if (recipe_no != jain.cell(row=i, column=1).value):
		recipe_no = jain.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == jain.cell(row=i, column=1).value):
			ingredients.append(jain.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For marathi recipes
csvfile = open('marathi.csv', 'wb')
writer = csv.writer(csvfile)
marathi = wb.get_sheet_by_name('Maharashtrian Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, marathi.max_row+1):
	if (recipe_no != marathi.cell(row=i, column=1).value):
		recipe_no = marathi.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == marathi.cell(row=i, column=1).value):
			ingredients.append(marathi.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For mughlai recipes
csvfile = open('mughlai.csv', 'wb')
writer = csv.writer(csvfile)
mughlai = wb.get_sheet_by_name('Mughlai Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, mughlai.max_row+1):
	if (recipe_no != mughlai.cell(row=i, column=1).value):
		recipe_no = mughlai.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == mughlai.cell(row=i, column=1).value):
			ingredients.append(mughlai.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)	

#For punjabi recipes
csvfile = open('punjabi.csv', 'wb')
writer = csv.writer(csvfile)
punjabi = wb.get_sheet_by_name('Punjabi Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, punjabi.max_row+1):
	if (recipe_no != punjabi.cell(row=i, column=1).value):
		recipe_no = punjabi.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == punjabi.cell(row=i, column=1).value):
			ingredients.append(punjabi.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For rajasthani recipes
csvfile = open('rajasthani.csv', 'wb')
writer = csv.writer(csvfile)
rajasthani = wb.get_sheet_by_name('Rajasthani Cuisine Recipes Ing')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, rajasthani.max_row+1):
	if (recipe_no != rajasthani.cell(row=i, column=1).value):
		recipe_no = rajasthani.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == rajasthani.cell(row=i, column=1).value):
			ingredients.append(rajasthani.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)

#For south_indian recipes
csvfile = open('south_indian.csv', 'wb')
writer = csv.writer(csvfile)
south_indian = wb.get_sheet_by_name('South Cuisine Recipes Ingred')
recipe_no = 'a'
ingredients = []
row = [recipe_no]

for i in range(2, south_indian.max_row+1):
	if (recipe_no != south_indian.cell(row=i, column=1).value):
		recipe_no = south_indian.cell(row=i, column=1).value
		row = [recipe_no]
		ingredients = []
		while (recipe_no == south_indian.cell(row=i, column=1).value):
			ingredients.append(south_indian.cell(row=i, column=2).value)
			i=i+1
		for k in range(0, len(ingredients)):
			row.append(ingredients[k])
		writer.writerow(row)